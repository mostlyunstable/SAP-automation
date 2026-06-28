"""Tests for sap_automation.utils.excel_reader — Excel input processing."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from sap_automation.core.exceptions import ExcelReadError
from sap_automation.utils.excel_reader import ExcelReader


@pytest.fixture
def simple_excel(tmp_path: Path) -> Path:
    """Create a simple Excel file with document numbers."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Document Number"])
    ws.append(["100001"])
    ws.append(["100002"])
    ws.append(["100003"])
    path = tmp_path / "simple.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def multi_column_excel(tmp_path: Path) -> Path:
    """Create an Excel file with multiple columns."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Document Number", "Plant", "Date"])
    ws.append(["100001", "1000", "2024-01-15"])
    ws.append(["100002", "2000", "2024-01-16"])
    path = tmp_path / "multi.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def empty_excel(tmp_path: Path) -> Path:
    """Create an empty Excel file (headers only)."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Document Number"])
    path = tmp_path / "empty.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def full_data_excel(tmp_path: Path) -> Path:
    """Create an Excel file without a document_number column."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Material", "Quantity", "Price"])
    ws.append(["MAT-001", 10, 25.50])
    ws.append(["MAT-002", 5, 100.00])
    path = tmp_path / "fulldata.xlsx"
    wb.save(path)
    return path


class TestExcelReaderInit:
    """Test initialization and validation."""

    def test_file_not_found(self) -> None:
        with pytest.raises(ExcelReadError, match="not found"):
            ExcelReader("/nonexistent/file.xlsx")

    def test_invalid_extension(self, tmp_path: Path) -> None:
        txt_file = tmp_path / "data.txt"
        txt_file.write_text("hello")
        with pytest.raises(ExcelReadError, match="Invalid file type"):
            ExcelReader(txt_file)

    def test_valid_xlsx(self, simple_excel: Path) -> None:
        reader = ExcelReader(simple_excel)
        assert reader.file_path == simple_excel

    def test_is_file_check(self, tmp_path: Path) -> None:
        with pytest.raises(ExcelReadError, match="not a file"):
            ExcelReader(tmp_path)


class TestExcelReaderRead:
    """Test reading Excel files."""

    def test_simple_list(self, simple_excel: Path) -> None:
        reader = ExcelReader(simple_excel)
        records = reader.read()
        assert len(records) == 3
        assert records[0]["document_number"] == "100001"
        assert records[1]["document_number"] == "100002"
        assert records[2]["document_number"] == "100003"

    def test_multi_column(self, multi_column_excel: Path) -> None:
        reader = ExcelReader(multi_column_excel)
        records = reader.read()
        assert len(records) == 2
        assert records[0]["Plant"] == "1000"
        assert records[0]["Date"] == "2024-01-15"

    def test_empty_file(self, empty_excel: Path) -> None:
        reader = ExcelReader(empty_excel)
        records = reader.read()
        assert len(records) == 0

    def test_full_data_pattern(self, full_data_excel: Path) -> None:
        reader = ExcelReader(
            full_data_excel, document_number_column="NonExistent"
        )
        records = reader.read()
        assert len(records) == 2
        # First non-null value used as document_number
        assert records[0]["document_number"] == "MAT-001"

    def test_headers_property(self, simple_excel: Path) -> None:
        reader = ExcelReader(simple_excel)
        reader.read()
        assert "Document Number" in reader.headers

    def test_custom_sheet_name(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Document Number"])
        ws.append(["100001"])
        wb.create_sheet("Other")
        path = tmp_path / "multi_sheet.xlsx"
        wb.save(path)

        reader = ExcelReader(path, sheet_name="Data")
        records = reader.read()
        assert len(records) == 1

    def test_invalid_sheet_name(self, simple_excel: Path) -> None:
        reader = ExcelReader(simple_excel, sheet_name="NonExistent")
        with pytest.raises(ExcelReadError, match="not found"):
            reader.read()


class TestExcelReaderValidate:
    """Test file validation without full read."""

    def test_valid_file(self, simple_excel: Path) -> None:
        reader = ExcelReader(simple_excel)
        issues = reader.validate()
        assert len(issues) == 0

    def test_missing_column(self, full_data_excel: Path) -> None:
        reader = ExcelReader(full_data_excel)
        issues = reader.validate()
        assert any("Document Number" in i for i in issues)


class TestExcelReaderEdgeCases:
    """Test edge cases and error handling."""

    def test_duplicate_document_numbers(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(["Document Number"])
        ws.append(["100001"])
        ws.append(["100001"])  # Duplicate
        ws.append(["100002"])
        path = tmp_path / "dupes.xlsx"
        wb.save(path)

        reader = ExcelReader(path)
        records = reader.read()
        assert len(records) == 3  # All records still returned

    def test_blank_rows_skipped(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.append(["Document Number"])
        ws.append(["100001"])
        ws.append([None])  # Blank row
        ws.append(["100002"])
        path = tmp_path / "blanks.xlsx"
        wb.save(path)

        reader = ExcelReader(path)
        records = reader.read()
        assert len(records) == 2

    def test_raw_data_preserved(self, simple_excel: Path) -> None:
        reader = ExcelReader(simple_excel)
        records = reader.read()
        assert "_raw" in records[0]
        assert "Document Number" in records[0]["_raw"]
