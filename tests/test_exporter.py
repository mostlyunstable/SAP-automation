"""Tests for sap_automation.utils.exporter — Report export (Excel, CSV, PDF)."""

from __future__ import annotations

import csv
from pathlib import Path

import pytest

from sap_automation.core.exceptions import ExportError
from sap_automation.utils.exporter import (
    ReportExporter,
    _sanitize_csv_value,
    sanitize_filename,
)


@pytest.fixture
def sample_data() -> list[dict[str, str]]:
    return [
        {"Name": "Item A", "Value": "100", "Status": "Active"},
        {"Name": "Item B", "Value": "200", "Status": "Inactive"},
    ]


@pytest.fixture
def exporter(tmp_path: Path) -> ReportExporter:
    return ReportExporter(output_dir=tmp_path, timestamp_in_name=False)


class TestSanitizeFilename:
    """Test filename sanitization."""

    def test_normal_name(self) -> None:
        assert sanitize_filename("report_123") == "report_123"

    def test_special_chars(self) -> None:
        result = sanitize_filename('report<>:"/\\|?*123')
        assert "<" not in result
        assert ">" not in result
        assert ":" not in result

    def test_empty_name(self) -> None:
        assert sanitize_filename("") == "report"

    def test_long_name(self) -> None:
        result = sanitize_filename("a" * 300, max_length=200)
        assert len(result) <= 200


class TestCsvSanitization:
    """Test CSV injection prevention."""

    def test_formula_prefix(self) -> None:
        assert _sanitize_csv_value("=SUM(A1)") == "'=SUM(A1)"

    def test_plus_prefix(self) -> None:
        assert _sanitize_csv_value("+cmd") == "'+cmd"

    def test_minus_prefix(self) -> None:
        assert _sanitize_csv_value("-cmd") == "'-cmd"

    def test_at_prefix(self) -> None:
        assert _sanitize_csv_value("@SUM") == "'@SUM"

    def test_normal_value(self) -> None:
        assert _sanitize_csv_value("hello") == "hello"

    def test_non_string(self) -> None:
        assert _sanitize_csv_value(123) == "123"

    def test_tab_in_value(self) -> None:
        result = _sanitize_csv_value("\tvalue")
        assert result == "'\tvalue"


class TestExcelExport:
    """Test Excel file export."""

    def test_export_creates_file(self, exporter: ReportExporter, sample_data: list[dict]) -> None:
        path = exporter.export_excel(sample_data, filename="test_report")
        assert path.exists()
        assert path.suffix == ".xlsx"

    def test_export_empty_data(self, exporter: ReportExporter) -> None:
        with pytest.raises(ExportError, match="No data"):
            exporter.export_excel([])

    def test_export_content(self, exporter: ReportExporter, sample_data: list[dict]) -> None:
        from openpyxl import load_workbook

        path = exporter.export_excel(sample_data, filename="content_test")
        wb = load_workbook(path)
        ws = wb.active
        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(2, 1).value == "Item A"
        assert ws.cell(3, 1).value == "Item B"
        wb.close()

    def test_custom_sheet_name(self, exporter: ReportExporter, sample_data: list[dict]) -> None:
        from openpyxl import load_workbook

        path = exporter.export_excel(sample_data, filename="sheet_test", sheet_name="MyData")
        wb = load_workbook(path)
        assert wb.active.title == "MyData"
        wb.close()


class TestCsvExport:
    """Test CSV file export."""

    def test_export_creates_file(self, exporter: ReportExporter, sample_data: list[dict]) -> None:
        path = exporter.export_csv(sample_data, filename="test_csv")
        assert path.exists()
        assert path.suffix == ".csv"

    def test_export_content(self, exporter: ReportExporter, sample_data: list[dict]) -> None:
        path = exporter.export_csv(sample_data, filename="csv_content")
        with open(path, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)
            assert rows[0] == ["Name", "Value", "Status"]
            assert rows[1][0] == "Item A"

    def test_csv_injection_prevented(self, exporter: ReportExporter) -> None:
        data = [{"Formula": "=SUM(A1)", "Normal": "text"}]
        path = exporter.export_csv(data, filename="injection_test")
        with open(path, encoding="utf-8-sig") as f:
            content = f.read()
            assert "'=SUM(A1)" in content


class TestMultiFormatExport:
    """Test multi-format export."""

    def test_export_xlsx_only(self, exporter: ReportExporter, sample_data: list[dict]) -> None:
        paths = exporter.export(sample_data, filename="multi", formats=["xlsx"])
        assert len(paths) == 1
        assert paths[0].suffix == ".xlsx"

    def test_export_all_formats(self, exporter: ReportExporter, sample_data: list[dict]) -> None:
        paths = exporter.export(sample_data, filename="all_fmt", formats=["all"])
        suffixes = {p.suffix for p in paths}
        assert ".xlsx" in suffixes
        assert ".csv" in suffixes
        # PDF should not be present since pywin32 is not available on macOS
        assert ".pdf" not in suffixes

    def test_export_pdf_fails_without_win32(self, exporter: ReportExporter, sample_data: list[dict]) -> None:
        from sap_automation.core.exceptions import ExportError

        with pytest.raises(ExportError, match="PDF export requires pywin32"):
            exporter.export_pdf(sample_data, filename="pdf_test")

    def test_export_unknown_format(self, exporter: ReportExporter, sample_data: list[dict]) -> None:
        paths = exporter.export(sample_data, filename="unknown", formats=["xyz"])
        assert len(paths) == 0

    def test_export_empty_data(self, exporter: ReportExporter) -> None:
        with pytest.raises(ExportError, match="No data"):
            exporter.export([], formats=["xlsx"])


class TestAtomicWrites:
    """Test that files are written atomically."""

    def test_no_temp_files_left(self, exporter: ReportExporter, sample_data: list[dict]) -> None:
        exporter.export_excel(sample_data, filename="atomic_test")
        # Check no .tmp files remain
        tmp_files = list(exporter.output_dir.glob("*.tmp"))
        assert len(tmp_files) == 0
