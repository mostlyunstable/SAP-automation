"""VA23 — Display Quotation transaction handler.

Implements the client's exact manual workflow:

1. Open VA23
2. Enter quotation number
3. Verify screen loaded
4. Trigger SAP's built-in export (List > Export > Spreadsheet)
5. Handle the Windows Save dialog
6. Verify the exported file
7. Return SAP to ready state

Field IDs are configurable and validated at runtime. If a field
cannot be found, the handler logs diagnostic information and
stops safely rather than continuing with partial data.

The exported file is the official SAP-generated Excel report —
the same file the client receives when performing the export
manually. This is NOT a Python-generated summary.
"""

from __future__ import annotations

import contextlib
import time
from pathlib import Path
from typing import Any

from ..core.base_transaction import BaseTransaction, TransactionResult
from ..core.config import Config
from ..core.exceptions import (
    ExportError,
    FieldError,
    SAPConnectionError,
    TransactionError,
    ValidationError,
)
from ..core.logger import get_logger

log = get_logger("transactions.va23")

# Default field IDs — may vary by SAP theme/version/screen layout
_DEFAULT_FIELD_IDS = {
    "input_vbeln": "wnd[0]/usr/ctxtRV45A-VBELN",
    "sold_to": "wnd[0]/usr/subHEADER/SUB1/RBHP-VERTR",
    "document_date": "wnd[0]/usr/subHEADER/SUB1/RBHP-AUDAT",
    "net_value": "wnd[0]/usr/subHEADER/SUB1/RBHP-NETWR",
    "currency": "wnd[0]/usr/subHEADER/SUB1/RBHP-WAERK",
}

# SAP export menu and dialog IDs — may vary by SAP version
_DEFAULT_EXPORT_IDS = {
    "export_menu": "wnd[0]/mbar/menu[0]/menu[3]/menu[1]",
    "format_dialog_ok": "wnd[1]/tbar[0]/btn[0]",
}

# Required fields for a valid VA23 record
_REQUIRED_FIELDS = {"document_number"}

# Maximum document number length for VA23
_MAX_DOC_NUM_LENGTH = 10


class VA23DisplayQuotation(BaseTransaction):
    """Handle VA23 (Display Quotation) with SAP-native export.

    Follows the client's exact manual workflow:
    - Opens VA23, enters quotation number
    - Verifies the screen loaded
    - Triggers SAP's built-in export (not a Python-generated file)
    - Handles the Windows Save dialog
    - Verifies the exported file exists and is valid
    - Returns SAP to ready state for the next quotation

    The exported file is the official SAP-generated Excel report —
    functionally identical to what the client receives manually.
    """

    def __init__(self, config: Config) -> None:
        super().__init__(config)
        self._field_ids = dict(_DEFAULT_FIELD_IDS)
        self._export_ids = dict(_DEFAULT_EXPORT_IDS)

        # Allow override from config
        field_overrides = config.get("transactions.va23.field_ids", {})
        if isinstance(field_overrides, dict):
            self._field_ids.update(field_overrides)

        export_overrides = config.get("transactions.va23.export_ids", {})
        if isinstance(export_overrides, dict):
            self._export_ids.update(export_overrides)

        # Export settings
        self._export_folder = config.get("paths.output_dir", "./output")
        self._export_format = config.get("export.format", "xlsx")

    @property
    def tcode(self) -> str:
        return "VA23"

    @property
    def description(self) -> str:
        return "Display Quotation"

    def validate_record(self, record: dict[str, Any]) -> None:
        """Validate a single input record.

        Checks:
            - document_number is present and non-empty
            - document_number is numeric
            - document_number is not too long (max 10 digits)

        Raises:
            ValidationError: If the record is invalid.
        """
        doc_num = record.get("document_number")

        if doc_num is None or (isinstance(doc_num, str) and not doc_num.strip()):
            raise ValidationError("Missing document_number")

        doc_str = str(doc_num).strip()
        if not doc_str:
            raise ValidationError("Missing document_number")

        if not doc_str.isdigit():
            raise ValidationError(
                f"Must be numeric, got: '{doc_str}'"
            )

        if len(doc_str) > _MAX_DOC_NUM_LENGTH:
            raise ValidationError(
                f"Document number too long: {len(doc_str)} chars "
                f"(max {_MAX_DOC_NUM_LENGTH})"
            )

        # Strip whitespace in-place for downstream use
        record["document_number"] = doc_str

    def pre_validate(self, session: Any, record: dict[str, Any]) -> None:
        """Pre-validate for VA23: verify input field exists."""
        super().pre_validate(session, record)

        field_id = self._field_ids["input_vbeln"]
        try:
            session.findById(field_id)
        except Exception as exc:
            raise TransactionError(
                f"VA23 input field not found: {field_id}",
                details={
                    "field_id": field_id,
                    "error": str(exc),
                    "suggestion": (
                        "The VA23 initial screen layout may differ. "
                        "Check SAP GUI theme, transaction variant, "
                        "and screen resolution."
                    ),
                },
            ) from exc

    def execute(self, session: Any, record: dict[str, Any]) -> dict[str, Any]:
        """Execute VA23 for a single quotation using SAP-native export.

        This method triggers SAP's built-in export, which produces the
        same official Excel report the client receives manually.

        Args:
            session: Active SAP GUI session.
            record: Must contain 'document_number'.

        Returns:
            Dictionary with exported file metadata.

        Raises:
            FieldError: If a required field cannot be found.
            TransactionError: If navigation or export fails.
            ExportError: If the file cannot be saved or verified.
        """
        doc_num = record["document_number"]
        log.info("Displaying quotation: %s", doc_num)

        # Step 1: Enter quotation number
        self._enter_document_number(session, doc_num)

        # Step 2: Handle popups after entering document
        self._handle_entry_popups(session)

        # Step 3: Verify we're on the quotation display screen
        self._verify_screen(session, doc_num)

        # Step 4: Trigger SAP export
        self._trigger_sap_export(session)

        # Step 5: Handle Windows Save dialog
        file_path = self._handle_save_dialog(doc_num)

        # Step 6: Close Excel window if SAP opened one
        self._close_excel_window()

        # Step 7: Verify exported file
        self._verify_exported_file(file_path)

        # Step 8: Return SAP to ready state
        self._return_to_ready(session)

        log.info(
            "Exported quotation %s: %s",
            doc_num,
            file_path,
        )

        return {
            "document_number": doc_num,
            "exported_file": str(file_path),
            "file_size": file_path.stat().st_size,
        }

    def run_with_session(
        self,
        sap: Any,
        record: dict[str, Any],
    ) -> TransactionResult:
        """Execute the transaction using an existing SAP connection.

        Overrides BaseTransaction to skip ReportExporter — SAP creates
        the primary Excel file directly. ReportExporter is only used
        for batch summaries, not for the primary export.

        Args:
            sap: Active SAPConnection instance.
            record: Input data dictionary.

        Returns:
            TransactionResult with execution details.
        """
        start_time = time.monotonic()
        result = TransactionResult(
            transaction_code=self.tcode,
            document_number=str(record.get("document_number", "")),
        )

        try:
            self.validate_record(record)

            if not sap.validate_session():
                raise SAPConnectionError(
                    "Session validation failed before transaction"
                )

            sap.open_transaction(self.tcode)

            self.pre_validate(sap.session, record)

            # Execute triggers SAP export and handles Save dialog
            export_data = self.execute(sap.session, record)
            result.data = export_data

            # SAP created the file — add to exported_files
            exported_path = Path(export_data.get("exported_file", ""))
            if exported_path.exists():
                result.exported_files = [exported_path]

            # Post-validate (check status bar, session health)
            self.post_validate(sap.session)

            result.success = True
            result.message = "Completed successfully"

        except ValidationError as exc:
            result.success = False
            result.message = f"Validation error: {exc}"
            result.error_type = "ValidationError"
            log.error(
                "Validation failed for %s: %s",
                result.document_number,
                exc,
            )

        except SAPConnectionError as exc:
            result.success = False
            result.message = f"Connection error: {exc}"
            result.error_type = "SAPConnectionError"
            log.error(
                "Connection error for %s: %s",
                result.document_number,
                exc,
            )

        except TransactionError as exc:
            result.success = False
            result.message = f"Transaction error: {exc}"
            result.error_type = "TransactionError"
            log.error(
                "Transaction failed for %s: %s",
                result.document_number,
                exc,
            )

        except ExportError as exc:
            result.success = False
            result.message = f"Export error: {exc}"
            result.error_type = "ExportError"
            log.error(
                "Export failed for %s: %s",
                result.document_number,
                exc,
            )

        except Exception as exc:
            result.success = False
            result.message = str(exc)
            result.error_type = type(exc).__name__
            log.error(
                "Unexpected error for %s: %s",
                result.document_number,
                exc,
                exc_info=True,
            )

        result.duration_seconds = time.monotonic() - start_time
        return result

    # -- SAP Navigation -------------------------------------------------------

    def _enter_document_number(self, session: Any, doc_num: str) -> None:
        """Enter the document number in the VA23 initial screen."""
        field_id = self._field_ids["input_vbeln"]
        try:
            input_field = session.findById(field_id)
            input_field.text = doc_num
            session.findById("wnd[0]").sendVKey(0)  # Enter
        except Exception as exc:
            raise FieldError(
                f"Failed to enter quotation number '{doc_num}'",
                details={
                    "field_id": field_id,
                    "error": str(exc),
                    "suggestion": (
                        f"Field '{field_id}' not found. "
                        "The VA23 screen layout may differ from expected. "
                        "Check SAP GUI theme and transaction variant."
                    ),
                },
            ) from exc

    def _handle_entry_popups(self, session: Any) -> None:
        """Handle popups that appear after entering document number."""
        with contextlib.suppress(Exception):
            session.findById("wnd[0]").sendVKey(0)  # Confirm info messages

    def _verify_screen(self, session: Any, doc_num: str) -> None:
        """Verify we're on the quotation display screen after navigation."""
        # Check status bar for errors
        try:
            status_bar = session.findById("wnd[0]/sbar")
            msg_type = str(status_bar.MessageType)
            msg_text = str(status_bar.Text)
            if msg_type == "E":
                raise TransactionError(
                    f"SAP error displaying quotation {doc_num}: {msg_text}"
                )
        except TransactionError:
            raise
        except Exception:  # noqa: S110 — status bar not accessible
            pass  # Status bar not accessible

        # Verify at least one header field exists (confirms screen loaded)
        header_field = self._field_ids.get("sold_to")
        if header_field:
            try:
                session.findById(header_field)
            except Exception:
                log.warning(
                    "Could not verify quotation screen — "
                    "header field '%s' not found. "
                    "Quotation %s may not exist or screen layout differs.",
                    header_field,
                    doc_num,
                )

    # -- SAP Export -----------------------------------------------------------

    def _trigger_sap_export(self, session: Any) -> None:
        """Trigger SAP's built-in export via the menu bar.

        Navigates: List > Export > Spreadsheet
        Then confirms the format selection dialog.
        """
        export_menu = self._export_ids.get(
            "export_menu", "wnd[0]/mbar/menu[0]/menu[3]/menu[1]"
        )

        try:
            log.info("Triggering SAP export: %s", export_menu)
            session.findById(export_menu).select()
            time.sleep(0.5)  # Wait for format dialog to appear
        except Exception as exc:
            raise TransactionError(
                "Failed to trigger SAP export",
                details={
                    "menu_path": export_menu,
                    "error": str(exc),
                    "suggestion": (
                        "The export menu path may differ in your SAP version. "
                        "Record a manual export to find the correct menu path. "
                        "Update transactions.va23.export_ids.export_menu in config."
                    ),
                },
            ) from exc

        # Confirm format selection dialog (accept default format)
        format_ok = self._export_ids.get(
            "format_dialog_ok", "wnd[1]/tbar[0]/btn[0]"
        )
        try:
            session.findById(format_ok).press()
            time.sleep(0.3)
            log.info("Format dialog confirmed")
        except Exception as exc:
            # Format dialog may not appear in all configurations
            log.warning("Could not confirm format dialog: %s", exc)

    def _handle_save_dialog(self, doc_num: str) -> Path:
        """Handle the Windows Save dialog that SAP opens.

        Uses Windows COM automation to fill in the filename and save.
        """
        from ..utils.windows_dialog import SaveDialogHandler

        handler = SaveDialogHandler()
        filename = f"VA23_{doc_num}.xlsx"

        return handler.save_file(
            filename=filename,
            folder=self._export_folder,
            delete_existing=True,
            overwrite_confirm=True,
        )

    def _close_excel_window(self) -> None:
        """Close the Excel window that SAP opens after xlsx export."""
        from ..utils.windows_dialog import SaveDialogHandler

        handler = SaveDialogHandler()
        handler.close_excel_window()

    def _verify_exported_file(self, file_path: Path) -> None:
        """Verify the exported file exists, is non-empty, and is a valid workbook."""
        if not file_path.exists():
            raise ExportError(
                f"Exported file not found: {file_path}",
                details={
                    "path": str(file_path),
                    "suggestion": (
                        "The file may not have been saved. "
                        "Check the target folder and SAP export status."
                    ),
                },
            )

        if not file_path.is_file():
            raise ExportError(
                f"Export path is not a file: {file_path}",
                details={"path": str(file_path)},
            )

        file_size = file_path.stat().st_size
        if file_size == 0:
            raise ExportError(
                f"Exported file is empty: {file_path}",
                details={
                    "path": str(file_path),
                    "suggestion": "The SAP export may have failed. Check SAP status bar.",
                },
            )

        # Verify workbook opens and has data
        try:
            from openpyxl import load_workbook

            wb = load_workbook(str(file_path), read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            if not sheet_names:
                raise ExportError(
                    f"Exported workbook has no sheets: {file_path}",
                    details={"path": str(file_path)},
                )

            log.info(
                "Verified export: %s (%d bytes, sheets: %s)",
                file_path,
                file_size,
                ", ".join(sheet_names),
            )
        except ExportError:
            raise
        except Exception as exc:
            raise ExportError(
                f"Cannot open exported workbook: {file_path}",
                details={
                    "path": str(file_path),
                    "error": str(exc),
                    "suggestion": (
                        "The file may be corrupted or locked. "
                        "Check if Excel has the file open."
                    ),
                },
            ) from exc

    def _return_to_ready(self, session: Any) -> None:
        """Return SAP to ready state for the next quotation.

        Presses Back to return to the VA23 initial screen.
        """
        try:
            session.findById("wnd[0]").sendVKey(3)  # Back (F3)
            time.sleep(0.3)
            log.debug("Returned SAP to ready state")
        except Exception as exc:
            log.warning("Could not return SAP to ready state: %s", exc)
