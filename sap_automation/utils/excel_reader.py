"""Excel input reader with two input patterns: document list or full data rows.

Features:
    - File existence and permission validation
    - Worksheet existence validation
    - Required column detection
    - Malformed record detection and reporting
    - Duplicate document number detection
    - Empty document number detection
    - Comprehensive validation report
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..core.exceptions import ExcelReadError
from ..core.logger import get_logger

log = get_logger("excel_reader")

_VALID_EXTENSIONS = {".xlsx", ".xlsm"}


class ExcelReader:
    """Read Excel files and yield records for transaction processing.

    Supports two input patterns:
        1. Document list: Single column of document numbers.
           The framework auto-maps to 'document_number'.
        2. Full data rows: Multiple columns. The column specified by
           ``document_number_column`` is mapped to 'document_number'.
           If not found, the first non-null value is used.

    Args:
        file_path: Path to the Excel file.
        document_number_column: Column header to use as document_number.
            If None, defaults to 'Document Number'.
        sheet_name: Sheet to read. If None, uses the active sheet.
    """

    def __init__(
        self,
        file_path: str | Path,
        document_number_column: str | None = None,
        sheet_name: str | None = None,
    ) -> None:
        self.file_path = Path(file_path)
        self.document_number_column = document_number_column or "Document Number"
        self.sheet_name = sheet_name
        self._headers: list[str] = []
        self._records: list[dict[str, Any]] = []

        self._validate_file()

    def _validate_file(self) -> None:
        """Validate the file exists, is readable, and has correct extension."""
        if not self.file_path.exists():
            raise ExcelReadError(
                f"File not found: {self.file_path}",
                details={
                    "path": str(self.file_path.resolve()),
                    "suggestion": "Check the file path and ensure the file exists.",
                },
            )

        if not self.file_path.is_file():
            raise ExcelReadError(
                f"not a file: {self.file_path}",
                details={
                    "path": str(self.file_path.resolve()),
                    "suggestion": "The path points to a directory, not a file.",
                },
            )

        if self.file_path.suffix.lower() not in _VALID_EXTENSIONS:
            raise ExcelReadError(
                f"Invalid file type: {self.file_path.suffix}. "
                f"Expected one of: {', '.join(sorted(_VALID_EXTENSIONS))}",
                details={
                    "path": str(self.file_path.resolve()),
                    "suffix": self.file_path.suffix,
                    "valid_extensions": sorted(_VALID_EXTENSIONS),
                },
            )

        try:
            with open(self.file_path, "rb") as f:
                f.read(1)
        except PermissionError as exc:
            raise ExcelReadError(
                f"File not readable (permission denied): {self.file_path}",
                details={
                    "path": str(self.file_path.resolve()),
                    "suggestion": "Check file permissions. Close the file if it is open in another application.",
                },
            ) from exc
        except OSError as exc:
            raise ExcelReadError(
                f"Cannot read file: {self.file_path}: {exc}",
                details={
                    "path": str(self.file_path.resolve()),
                },
            ) from exc

    @property
    def headers(self) -> list[str]:
        """Column headers from the last read operation."""
        return list(self._headers)

    def read(self) -> list[dict[str, Any]]:
        """Read all records from the Excel file.

        Returns:
            List of record dictionaries. Each record contains:
            - 'document_number': The mapped document number
            - All original columns preserved
            - '_raw': Original row data as dict

        Raises:
            ExcelReadError: If file cannot be read or parsed.
        """
        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
        except Exception as exc:
            raise ExcelReadError(
                f"Cannot open Excel file: {exc}",
                details={
                    "path": str(self.file_path.resolve()),
                    "suggestion": (
                        "Ensure the file is a valid Excel file and is not corrupted. "
                        "If the file is open in Excel, close it first."
                    ),
                },
            ) from exc

        try:
            ws = self._get_worksheet(wb)
            self._headers, self._records = self._parse_sheet(ws)
        finally:
            wb.close()

        issues = self._validate_records()
        if issues:
            for issue in issues:
                log.warning("Input validation: %s", issue)

        log.info(
            "Read %d records from %s (%d issues)",
            len(self._records),
            self.file_path.name,
            len(issues),
        )
        return list(self._records)

    def _get_worksheet(self, wb: Any) -> Worksheet:
        """Get the target worksheet."""
        if self.sheet_name:
            if self.sheet_name not in wb.sheetnames:
                raise ExcelReadError(
                    f"Sheet '{self.sheet_name}' not found.",
                    details={
                        "requested_sheet": self.sheet_name,
                        "available_sheets": wb.sheetnames,
                        "suggestion": (
                            f"Available sheets: {', '.join(wb.sheetnames)}. "
                            "Check the sheet name in your config."
                        ),
                    },
                )
            return wb[self.sheet_name]
        return wb.active

    def _parse_sheet(
        self, ws: Worksheet
    ) -> tuple[list[str], list[dict[str, Any]]]:
        """Parse worksheet into headers and records."""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], []

        headers = [str(h) if h is not None else f"Col{i}" for i, h in enumerate(rows[0])]

        # Find document_number column index
        doc_col_idx = self._find_document_number_column(headers)

        records: list[dict[str, Any]] = []
        skipped_blank = 0
        for row in rows[1:]:
            # Skip blank rows
            if all(cell is None for cell in row):
                skipped_blank += 1
                continue

            record: dict[str, Any] = {}
            raw: dict[str, Any] = {}

            for i, header in enumerate(headers):
                value = row[i] if i < len(row) else None
                # Normalize floats: int-like floats become ints
                if isinstance(value, float) and value == int(value):
                    value = int(value)
                record[header] = value
                raw[header] = value

            if doc_col_idx >= 0 and doc_col_idx < len(row):
                doc_value = row[doc_col_idx]
                record["document_number"] = str(doc_value) if doc_value is not None else ""
            else:
                # Fallback: use first non-null value
                for cell in row:
                    if cell is not None:
                        record["document_number"] = str(cell)
                        break
                else:
                    record["document_number"] = ""

            record["_raw"] = raw
            records.append(record)

        if skipped_blank > 0:
            log.info("Skipped %d blank rows", skipped_blank)

        return headers, records

    def _find_document_number_column(self, headers: list[str]) -> int:
        """Find the index of the document_number column."""
        for i, header in enumerate(headers):
            if header.lower() == self.document_number_column.lower():
                return i
        return -1

    def _validate_records(self) -> list[str]:
        """Validate parsed records for common issues.

        Returns:
            List of warning strings. Empty if no issues found.
        """
        issues: list[str] = []
        seen_doc_nums: dict[str, int] = {}
        empty_count = 0

        for i, record in enumerate(self._records):
            doc_num = record.get("document_number", "")

            if not doc_num:
                empty_count += 1
                issues.append(
                    f"Row {i + 2}: Empty document_number"
                )
                continue

            if doc_num in seen_doc_nums:
                issues.append(
                    f"Row {i + 2}: Duplicate document_number '{doc_num}' "
                    f"(first at row {seen_doc_nums[doc_num]})"
                )
            else:
                seen_doc_nums[doc_num] = i + 2

        if empty_count > 0:
            issues.append(
                f"{empty_count} row(s) have empty document_number "
                f"and will be processed with empty document number"
            )

        return issues

    def validate(self) -> list[str]:
        """Validate the file structure without full read.

        Returns:
            List of issue strings. Empty if no issues found.
        """
        issues: list[str] = []

        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
        except Exception as exc:
            return [f"Cannot open file: {exc}"]

        try:
            ws = self._get_worksheet(wb)
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return ["File is empty"]

            headers = [str(h) if h is not None else "" for h in rows[0]]

            doc_col_found = any(
                h.lower() == self.document_number_column.lower() for h in headers
            )
            if not doc_col_found:
                issues.append(
                    f"Column '{self.document_number_column}' not found. "
                    f"Available: {', '.join(headers)}"
                )

            # Count data rows
            data_rows = len(rows) - 1
            if data_rows == 0:
                issues.append("File has headers but no data rows")

        finally:
            wb.close()

        return issues
